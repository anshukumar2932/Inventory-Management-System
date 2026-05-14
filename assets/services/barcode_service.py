import hashlib
import io
import time

import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage

from assets.models import Asset


class BarcodeService:

    @staticmethod
    def generate_barcode_string(raw: str) -> str:
        return hashlib.sha256(raw.encode()).hexdigest()[:20].upper()

    @staticmethod
    def generate_barcode_image(code: str):
        try:
            writer = ImageWriter()
            writer.set_options({
                "module_width": 0.2,
                "module_height": 15,
                "font_size": 8,
                "text_distance": 2,
                "quiet_zone": 2,
            })
            barcode_cls = barcode.get_barcode_class("code128")
            bar = barcode_cls(code, writer=writer)
            buf = io.BytesIO()
            bar.write(buf)
            buf.seek(0)
            return buf
        except Exception:
            return None

    @staticmethod
    def auto_generate_barcode(asset_code: str, asset_name: str = "") -> str:
        raw = f"{asset_code}{asset_name}{time.time()}"
        return BarcodeService.generate_barcode_string(raw)

    @staticmethod
    def generate_barcode_excel(asset_ids: list) -> bytes:
        assets = Asset.objects.filter(id__in=asset_ids).order_by('id')
        if not assets.exists():
            return b""

        wb = Workbook()
        ws = wb.active
        ws.title = "Barcodes"
        ws.append(["Asset Code", "Barcode", "Barcode Image"])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25

        for i, asset in enumerate(assets, start=2):
            ws.cell(row=i, column=1, value=asset.asset_code)
            ws.cell(row=i, column=2, value=asset.barcode)
            buf = BarcodeService.generate_barcode_image(asset.barcode)
            if buf:
                img = XlImage(buf)
                img.width = 120
                img.height = 40
                ws.add_image(img, f"C{i}")
                ws.row_dimensions[i].height = 45

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
