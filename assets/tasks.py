import io
import time

from celery import shared_task
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage

from assets.models import Asset
from helper.barcode_generator import generate_barcode_image


@shared_task(bind=True, max_retries=2)
def generate_barcode_excel_task(self, asset_ids):
    assets = Asset.objects.filter(id__in=asset_ids).order_by('id')
    if not assets.exists():
        return "No assets found"

    wb = Workbook()
    ws = wb.active
    ws.title = "Barcodes"
    ws.append(["Asset Code", "Barcode", "Barcode Image"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25

    for i, asset in enumerate(assets, start=2):
        ws.cell(row=i, column=1, value=asset.asset_code)
        ws.cell(row=i, column=2, value=asset.barcode)
        buf = generate_barcode_image(asset.barcode)
        if buf:
            img = XlImage(buf)
            img.width = 120
            img.height = 40
            ws.add_image(img, f"C{i}")
            ws.row_dimensions[i].height = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
