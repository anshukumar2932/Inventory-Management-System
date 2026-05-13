# API views that power the inventory management system
# Each ViewSet maps to a model and provides CRUD + search/filter endpoints
from collections import Counter
from datetime import timedelta
from django.db.models import Count, Sum
from django.utils import timezone
from helper.barcode_generator import barcode_generator, generate_barcode_image
import io
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import BasePermission
from rest_framework.response import Response
from accounts.models import Department
from repairs.models import RepairTicket
from procurement.models import ProcurementRequest
from .models import Asset, Category, Location, Document, ServiceType, AssetService
from vendors.models import Vendor, Service, ClientCompany
from .serializers import (
    AssetSerializer,
    AssetListSerializer,
    AssetDetailSerializer,
    CategorySerializer,
    LocationSerializer,
    VendorSerializer,
    DocumentSerializer,
    ServiceTypeSerializer,
    AssetServiceSerializer,
)
from rest_framework import viewsets
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from .response import success_response, error_response

# ── Role-based permissions ──────────────────────────────────
# USER         — read only
# MANAGER      — read + create/update assets
# DEPT_ADMIN   — read + create/update/delete within department, manage dept users
# SUPER_ADMIN  — full system access, manage departments, all users

class IsAuth(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated


class IsManagerOrAbove(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ("MANAGER", "DEPARTMENT_ADMIN", "SUPER_ADMIN")


class IsDeptAdminOrAbove(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ("DEPARTMENT_ADMIN", "SUPER_ADMIN")


class IsSuperAdmin(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == "SUPER_ADMIN"
     

# ── Dashboard Stats ──────────────────────────────────────────
# Aggregated numbers shown on the main dashboard page.
# Hits the database with efficient COUNT queries.

@api_view(["GET"])
@permission_classes([IsAuth])
def dashboard_stats(request):
    user = request.user
    a_qs = Asset.objects.all()
    r_qs = RepairTicket.objects.all()
    p_qs = ProcurementRequest.objects.all()

    if user.role != "SUPER_ADMIN" and user.department:
        a_qs = a_qs.filter(department=user.department)
        r_qs = r_qs.filter(asset__department=user.department)
        p_qs = p_qs.filter(department=user.department)
    elif user.role != "SUPER_ADMIN":
        a_qs = Asset.objects.none()
        r_qs = RepairTicket.objects.none()
        p_qs = ProcurementRequest.objects.none()

    total = a_qs.count()
    active = a_qs.filter(status="ACTIVE").count()
    repair = a_qs.filter(status="REPAIR").count()
    missing = a_qs.filter(status="MISSING").count()
    retired = a_qs.filter(status="RETIRED").count()
    blocked = a_qs.filter(status="BLOCKED").count()

    pending_procurements = p_qs.filter(approval_status="PENDING").count()

    cats = (
        a_qs.values("category__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    assets_by_category = [{"name": c["category__name"], "count": c["count"]} for c in cats]

    depts = (
        a_qs.values("department__department_name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    assets_by_department = [{"name": d["department__department_name"], "count": d["count"]} for d in depts]

    months = []
    now = timezone.now()
    for i in range(11, -1, -1):
        dt = now - timedelta(days=30 * i)
        ym = dt.strftime("%Y-%m")
        months.append(ym)
    monthly_raw = (
        a_qs.extra(select={"ym": "to_char(created_at, 'YYYY-MM')"})
        .values("ym")
        .annotate(count=Count("id"))
        .order_by("ym")
    )
    monthly_map = {m["ym"]: m["count"] for m in monthly_raw}
    monthly_additions = [{"month": m, "count": monthly_map.get(m, 0)} for m in months]

    repair_total = r_qs.aggregate(s=Sum("repair_cost"))["s"] or 0
    repair_analytics = {
        "open_tickets": r_qs.exclude(status__iexact="closed").count(),
        "closed_tickets": r_qs.filter(status__iexact="closed").count(),
        "total_cost": float(repair_total),
    }

    audit_trends = []

    return Response({
        "total_assets": total,
        "active_assets": active,
        "repair_assets": repair,
        "missing_assets": missing,
        "retired_assets": retired,
        "blocked_assets": blocked,
        "pending_procurements": pending_procurements,
        "assets_by_category": assets_by_category,
        "assets_by_department": assets_by_department,
        "monthly_additions": monthly_additions,
        "repair_analytics": repair_analytics,
        "audit_trends": audit_trends,
    })


# ── Category CRUD ────────────────────────────────────────────
# Categories are simple name-only resources.
# SearchFilter lets the frontend do ?search=electronics.

class CategoryViewSet(viewsets.ModelViewSet):

    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name']

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsSuperAdmin()]
        return [IsAuth()]

    @action(detail=False, methods=['POST'], permission_classes=[IsSuperAdmin])
    def add(self, request):
        name = request.data.get("name")
        if not name:
            return Response(
                {"error": "Name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        category, created = Category.objects.get_or_create(name=name)
        serializer = self.get_serializer(category)
        return Response(
            {"created": created, "data": serializer.data},
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['PATCH'], permission_classes=[IsSuperAdmin])
    def update_category(self, request):
        name = request.data.get("name")
        if not name:
            return Response(
                {"error": "Name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            category = Category.objects.get(name=name)
        except Category.DoesNotExist:
            return Response(
                {"error": "Category not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = self.get_serializer(category, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"data": serializer.data})

    @action(detail=False, methods=['DELETE'], permission_classes=[IsSuperAdmin])
    def remove_category(self, request):
        name = request.data.get("name")
        if not name:
            return Response(
                {"error": "Name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            category = Category.objects.get(name=name)
        except Category.DoesNotExist:
            return Response(
                {"error": "Category not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Location CRUD ────────────────────────────────────────────
# Physical locations where assets are kept.

class LocationViewSet(viewsets.ModelViewSet):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    filter_backends = [SearchFilter, OrderingFilter]

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsSuperAdmin()]
        return [IsAuth()]
    search_fields = ['name']

    @action(
        detail=False,
        methods=['POST'],
        permission_classes=[IsSuperAdmin]
    )
    def add(self, request):
        name = request.data.get("name")
        if not name:
            return Response(
                {"error": "Name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        location, created = Location.objects.get_or_create(name=name)
        serializer = self.get_serializer(location)
        return Response(
            {"created": created, "data": serializer.data},
            status=status.HTTP_201_CREATED
        )

    @action(
        detail=False,
        methods=['PATCH'],
        permission_classes=[IsSuperAdmin]
    )
    def update_location(self, request):
        name = request.data.get("name")
        if not name:
            return Response(
                {"error": "Name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            location = Location.objects.get(name=name)
        except Location.DoesNotExist:
            return Response(
                {"error": "Location not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = self.get_serializer(location, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"data": serializer.data})


# ── Document Upload/Download ────────────────────────────────

class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuth]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        doc = Document.objects.create(
            file_name=file.name,
            content_type=file.content_type,
            file_size=file.size,
            file_data=file.read(),
        )
        serializer = self.get_serializer(doc)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['GET'])
    def download(self, request, pk=None):
        doc = self.get_object()
        return HttpResponse(doc.file_data, content_type=doc.content_type)


# ── Service Types ────────────────────────────────────────────

class ServiceTypeViewSet(viewsets.ModelViewSet):
    queryset = ServiceType.objects.all()
    serializer_class = ServiceTypeSerializer
    permission_classes = [IsAuth]

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsSuperAdmin()]
        return [IsAuth()]


# ── Asset Services ───────────────────────────────────────────

class AssetServiceViewSet(viewsets.ModelViewSet):
    queryset = AssetService.objects.none()
    serializer_class = AssetServiceSerializer
    permission_classes = [IsAuth]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['asset', 'service_type', 'status', 'provider']

    def get_queryset(self):
        qs = AssetService.objects.select_related('service_type', 'provider', 'asset')
        user = self.request.user
        if user.role == "SUPER_ADMIN":
            return qs.all()
        if user.department:
            return qs.filter(asset__department=user.department)
        return qs.none()

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsManagerOrAbove()]
        return [IsAuth()]


# ── Vendor CRUD ──────────────────────────────────────────────
# Vendors supply the assets. Searchable by vendor_name.

class VendorViewSet(viewsets.ModelViewSet):

    serializer_class = VendorSerializer
    permission_classes = [IsAuth]

    filter_backends = [SearchFilter, OrderingFilter]

    search_fields = [
        'vendor_name',
        'vendor_code',
        'contact_person',
        'email',
        'phone',
        'gst_number',
    ]

    ordering_fields = [
        'vendor_name',
        'vendor_code',
        'created_at',
        'updated_at',
        'rating',
    ]

    ordering = ['vendor_name']

    def get_queryset(self):

        queryset = Vendor.objects.filter(
            is_deleted=False
        ).select_related(
            'vendor_category'
        ).prefetch_related(
            'services',
            'supported_categories',
            'served_companies',
            'bank_accounts',
            'contacts',
        )

        status_param = self.request.query_params.get('status')

        if status_param:
            queryset = queryset.filter(status=status_param)

        category = self.request.query_params.get('category')

        if category:
            queryset = queryset.filter(
                vendor_category__id=category
            )

        return queryset

    def generate_vendor_code(self):

        last_vendor = Vendor.objects.order_by('-id').first()

        if not last_vendor:
            return "VEND0001"

        try:
            last_number = int(
                last_vendor.vendor_code.replace('VEND', '')
            )
        except:
            last_number = last_vendor.id

        return f"VEND{last_number + 1:04d}"

    @action(
        detail=False,
        methods=['POST'],
        permission_classes=[IsAuth]
    )
    def add(self, request):

        vendor_name = request.data.get("vendor_name")

        if not vendor_name:
            return Response(
                {
                    "error": "vendor_name is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        vendor_category = None

        category_id = request.data.get("vendor_category")

        if category_id:
            try:
                vendor_category = VendorCategory.objects.get(
                    id=category_id
                )
            except VendorCategory.DoesNotExist:
                return Response(
                    {
                        "error": "Invalid vendor category"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

        defaults = {

            "vendor_code": self.generate_vendor_code(),

            "contact_person": request.data.get(
                "contact_person",
                ""
            ),

            "email": request.data.get(
                "email",
                f"{vendor_name.replace(' ', '_').lower()}@vendor.com"
            ),

            "phone": request.data.get(
                "phone",
                ""
            ),

            "alternate_phone": request.data.get(
                "alternate_phone",
                ""
            ),

            "address": request.data.get(
                "address",
                ""
            ),

            "gst_number": request.data.get(
                "gst_number",
                ""
            ),

            "pan_number": request.data.get(
                "pan_number",
                ""
            ),

            "remarks": request.data.get(
                "remarks",
                ""
            ),

            "status": request.data.get(
                "status",
                "PENDING"
            ),

            "rating": request.data.get(
                "rating",
                0.0
            ),

            "vendor_category": vendor_category,
        }

        vendor, created = Vendor.objects.get_or_create(
            vendor_name=vendor_name,
            defaults=defaults
        )

        if created:
            for s in request.data.get("services", "").split(","):
                s = s.strip()
                if s:
                    service, _ = Service.objects.get_or_create(service_name=s)
                    vendor.services.add(service)

            for c in request.data.get("supported_categories", "").split(","):
                c = c.strip()
                if c:
                    cat, _ = Category.objects.get_or_create(name=c)
                    vendor.supported_categories.add(cat)

            for co in request.data.get("served_companies", "").split(","):
                co = co.strip()
                if co:
                    company, _ = ClientCompany.objects.get_or_create(company_name=co)
                    vendor.served_companies.add(company)

        serializer = self.get_serializer(vendor)

        return Response(
            {
                "created": created,
                "data": serializer.data
            },
            status=status.HTTP_201_CREATED
        )

    @action(
        detail=True,
        methods=['DELETE'],
        permission_classes=[IsAuth]
    )
    def soft_delete(self, request, pk=None):

        vendor = self.get_object()

        vendor.is_deleted = True
        vendor.save()

        return Response(
            {
                "message": "Vendor deleted successfully"
            },
            status=status.HTTP_200_OK
        )

# ── Asset CRUD ───────────────────────────────────────────────
# The main resource. Everything else supports this.
# Uses select_related to avoid N+1 queries on FK fields.
# Uses different serializers for list vs detail for performance.

class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.none()
    serializer_class = AssetSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['asset_code', 'asset_name', 'brand', 'model_name', 'serial_number']
    filterset_fields = ["procurement_request", "status", "approval_status", "category", "department", "location"]

    def get_queryset(self):
        qs = Asset.objects.select_related("category", "location", "department", "vendor")
        user = self.request.user
        if user.role == "SUPER_ADMIN":
            return qs.all()
        if user.department:
            return qs.filter(department=user.department)
        return qs.none()

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsManagerOrAbove()]
        return [IsAuth()]

    def perform_update(self, serializer):
        user = self.request.user
        if user.role == "MANAGER" and 'status' in serializer.validated_data:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Managers cannot change asset status. Only SUPER_ADMIN or DEPARTMENT_ADMIN can.")
        serializer.save()

    def get_serializer_class(self):
        if self.action == "list":
            return AssetListSerializer  # Lightweight — only table columns
        if self.action == "retrieve":
            return AssetDetailSerializer  # Full detail with resolved names
        return AssetSerializer  # Full fields for writes

    @action(
        detail=False,
        methods=['POST'],
        permission_classes=[IsManagerOrAbove]
    )
    def add(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return success_response(
            message="Asset created successfully",
            http_status=status.HTTP_201_CREATED
        )

    @action(
        detail=False,
        methods=['PATCH'],
        permission_classes=[IsSuperAdmin]
    )
    def update_asset(self, request):
        asset_name = request.data.get("asset_name")
        if not asset_name:
            return error_response("asset_name is required")
        try:
            asset = Asset.objects.get(asset_name=asset_name)
        except Asset.DoesNotExist:
            return error_response(
                "Asset not found",
                http_status=status.HTTP_404_NOT_FOUND
            )
        serializer = self.get_serializer(asset, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return success_response(message="Asset updated successfully")
    
    @action(
        detail=False,
        methods=['POST','GET'],
        permission_classes=[IsManagerOrAbove]
    )
    def bulk_upload(self, request):
        # GET returns a template Excel file for download
        if request.method == "GET":
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            headers_row = [
                "asset_code", "asset_name", "category", "brand", "model_name",
                "location", "department", "serial_number", "manufacturer",
                "barcode", "model_detail", "invoice_number", "status",
                "vendor"
            ]
            ws.append(headers_row)
            ws.append(["AST001", "Laptop Dell XPS", "Electronics", "Dell", "XPS 15",
                        "Main Office", "IT", "SN123456", "Dell Inc.",
                        "", "", "", "ACTIVE", "TechVendor"])
            for col_idx in range(1, len(headers_row) + 1):
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="bulk_upload_template.xlsx"'
            return response

        # POST processes an uploaded file — CSV or Excel
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        ext = file.name.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            df = pd.read_csv(file)
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(file, engine="openpyxl" if ext == "xlsx" else "xlrd")
        else:
            return Response({"error": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)

        # Normalize column names — strip whitespace, lowercase, replace spaces
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        df = df.where(pd.notna(df), None)

        errors = []
        created_assets = []

        def val(v, default=""):
            if v is None or (isinstance(v, float) and __import__('math').isnan(v)):
                return default
            return v

        for idx, row in df.iterrows():
            row_errors = []

            try:
                category_name = row.get("category")
                location_name = row.get("location")
                department_name = row.get("department")
                vendor_name = row.get("vendor")

                category_name = str(category_name).strip() if category_name else None
                location_name = str(location_name).strip() if location_name else None
                department_name = str(department_name).strip() if department_name else None
                vendor_name = str(vendor_name).strip() if vendor_name else None

                if not category_name:
                    row_errors.append("category is required")
                if not location_name:
                    row_errors.append("location is required")
                if not department_name:
                    row_errors.append("department is required")

                if row_errors:
                    errors.append({"row": idx + 2, "message": "; ".join(row_errors)})
                    continue

                # Find or create the related objects by name
                category = Category.objects.filter(name__iexact=category_name).first()
                if not category:
                    category = Category.objects.create(name=category_name)

                location = Location.objects.filter(name__iexact=location_name).first()
                if not location:
                    location = Location.objects.create(name=location_name, building="", floor="", room="")

                department = Department.objects.filter(department_name__iexact=department_name).first()
                if not department:
                    department = Department.objects.create(department_name=department_name)

                vendor = None
                if vendor_name:
                    vendor = Vendor.objects.filter(vendor_name__iexact=vendor_name).first()
                    if not vendor:
                        vendor = Vendor.objects.create(
                            vendor_name=vendor_name,
                            contact_person="",
                            email=f"vendor_{vendor_name.replace(' ', '_').lower()}@temp.com",
                            phone="",
                            address="",
                        )

                barcode_val = val(row.get("barcode"))
                if not barcode_val:
                    raw = f"{val(row.get('asset_code'))}{val(row.get('asset_name'))}{idx}{__import__('time').time()}"
                    barcode_val = barcode_generator(raw)

                data = {
                    "asset_code": val(row.get("asset_code")),
                    "barcode": barcode_val,
                    "asset_name": val(row.get("asset_name")),
                    "category": category.id,
                    "brand": val(row.get("brand")),
                    "model_name": val(row.get("model_name")),
                    "location": location.id,
                    "department": department.id,
                    "serial_number": str(val(row.get("serial_number", ""))),
                    "model_detail": val(row.get("model_detail")),
                    "manufacturer": val(row.get("manufacturer")),
                    "invoice_number": val(row.get("invoice_number")),
                    "status": val(row.get("status"), "ACTIVE"),
                    "vendor": vendor.id if vendor else None,
                }

                serializer = AssetSerializer(data=data)
                if serializer.is_valid():
                    asset = serializer.save()
                    created_assets.append(asset)
                else:
                    msg = "; ".join(f"{k}: {', '.join(v)}" for k, v in serializer.errors.items())
                    errors.append({"row": idx + 2, "message": msg})

            except Exception as e:
                errors.append({"row": idx + 2, "message": str(e)})

        if not created_assets and errors:
            return Response({"error": "No assets created", "details": errors}, status=status.HTTP_400_BAD_REQUEST)

        # Generate a barcode-spreadsheet for the newly created assets
        wb = Workbook()
        ws = wb.active
        ws.title = "Barcodes"
        ws.append(["Asset Code", "Barcode", "Barcode Image"])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25

        for i, asset in enumerate(created_assets, start=2):
            ws.cell(row=i, column=1, value=asset.asset_code)
            ws.cell(row=i, column=2, value=asset.barcode)
            buf = generate_barcode_image(asset.barcode)
            if buf:
                img = XlImage(buf)
                img.width = 120
                img.height = 40
                ws.add_image(img, f"C{i}")
                ws.row_dimensions[i].height = 45

        if errors:
            ws2 = wb.create_sheet("Errors")
            ws2.append(["Row", "Error"])
            ws2.column_dimensions["A"].width = 10
            ws2.column_dimensions["B"].width = 60
            for err in errors:
                ws2.append([err["row"], err["message"]])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="barcodes.xlsx"'
        return response

    @action(
        detail=False,
        methods=['POST'],
        permission_classes=[IsAuth]
    )
    def scan(self, request):

        barcode = request.data.get("barcode")

        if not barcode:
            return Response(
                {"error": "barcode required"},
                status=400
            )

        try:
            asset = Asset.objects.select_related(
                "category",
                "location",
                "department"
            ).get(barcode=barcode)

        except Asset.DoesNotExist:
            return Response(
                {"error": "Asset not found"},
                status=404
            )

        serializer = AssetDetailSerializer(asset)

        return Response({
            "found": True,
            "asset": serializer.data
        })
