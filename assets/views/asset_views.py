import io

from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from assets.models import Asset
from assets.serializers import AssetSerializer, AssetListSerializer, AssetDetailSerializer
from assets.response import success_response, error_response
from assets.services.barcode_service import BarcodeService
from assets.services.upload_service import BulkUploadService
from assets.tasks import generate_barcode_excel_task
from procurement.models import ProcurementRequest

from .permissions import IsAuth, IsManagerOrAbove, IsSuperAdmin


class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.none()
    serializer_class = AssetSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['asset_code', 'asset_name', 'brand', 'model_name', 'serial_number']
    filterset_fields = ["procurement_request", "status", "approval_status", "category", "department", "location"]

    def _get_or_create_procurement_request(self, serializer):
        procurement = serializer.validated_data.get("procurement_request")
        auto_created = False

        if procurement is None:
            procurement = ProcurementRequest.objects.create(
                department=serializer.validated_data["department"],
                requested_by=self.request.user,
                remarks=self.request.data.get("remarks", "") or "Auto-created from asset intake",
            )
            auto_created = True

        return procurement, auto_created

    def perform_create(self, serializer):
        procurement, auto_created = self._get_or_create_procurement_request(serializer)
        instance = serializer.save(
            procurement_request=procurement,
            approval_status='PENDING',
            status='PROCUREMENT',
        )
        if auto_created:
            from assets.services.send_email import send_procurement_approval_email
            send_procurement_approval_email(procurement)

    def get_queryset(self):
        from assets.services.asset_service import AssetService
        include_unapproved = bool(self.request.query_params.get("procurement_request"))
        return AssetService.get_scoped_asset_qs(
            self.request.user,
            include_unapproved=include_unapproved,
        ).prefetch_related("asset_services__service_type")

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsManagerOrAbove()]
        return [IsAuth()]

    def perform_update(self, serializer):
        user = self.request.user
        if user.role == "MANAGER" and 'status' in serializer.validated_data:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Managers cannot change asset status")
        serializer.save()

    def get_serializer_class(self):
        if self.action == "list":
            return AssetListSerializer
        if self.action == "retrieve":
            return AssetDetailSerializer
        return AssetSerializer

    @action(detail=False, methods=['POST'], permission_classes=[IsManagerOrAbove])
    def add(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        procurement, auto_created = self._get_or_create_procurement_request(serializer)
        serializer.save(
            procurement_request=procurement,
            approval_status='PENDING',
            status='PROCUREMENT',
        )
        if auto_created:
            from assets.services.send_email import send_procurement_approval_email
            send_procurement_approval_email(procurement)
        return success_response(message="Asset created successfully", http_status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['PATCH'], permission_classes=[IsSuperAdmin])
    def update_asset(self, request):
        asset_name = request.data.get("asset_name")
        if not asset_name:
            return error_response("asset_name is required")
        try:
            asset = Asset.objects.get(asset_name=asset_name)
        except Asset.DoesNotExist:
            return error_response("Asset not found", http_status=status.HTTP_404_NOT_FOUND)
        serializer = self.get_serializer(asset, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return success_response(message="Asset updated successfully")

    @action(detail=False, methods=['POST', 'GET'], permission_classes=[IsManagerOrAbove])
    def bulk_upload(self, request):
        if request.method == "GET":
            return BulkUploadService.generate_template()

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            created_assets, errors = BulkUploadService.process_upload(file)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if not created_assets and errors:
            return Response({"error": "No assets created", "details": errors}, status=status.HTTP_400_BAD_REQUEST)

        if created_assets:
            from collections import defaultdict
            from assets.services.send_email import send_new_asset_email
            dept_groups = defaultdict(list)
            for asset in created_assets:
                if asset.department:
                    dept_groups[asset.department].append(asset.id)
            for dept, asset_ids in dept_groups.items():
                send_new_asset_email(Asset.objects.filter(pk__in=asset_ids), dept)

        wb = self._build_barcode_excel(created_assets, errors)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="barcodes.xlsx"'
        return response

    def _build_barcode_excel(self, created_assets, errors):
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage

        wb = Workbook()
        ws = wb.active
        ws.title = "Barcodes"
        ws.append(["Asset Code", "Barcode", "Barcode Image"])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25

        for i, asset in enumerate(created_assets, start=2):
            ws.cell(row=i, column=1, value=asset.asset_code)
            ws.cell(row=i, column=2, value=asset.barcode)
            buf = BarcodeService.generate_barcode_image(asset.barcode)
            if buf:
                img = XlImage(buf)
                img.width = 120
                img.height = 40
                ws.add_image(img, f"C{i}")
                ws.row_dimensions[i].height = 45

        if errors:
            ws2 = wb.create_sheet("Errors")
            ws2.append(["Row", "Error"])
            ws2.column_dimensions["A"].width = 10
            ws2.column_dimensions["B"].width = 60
            for err in errors:
                ws2.append([err["row"], err["message"]])

        return wb

    @action(detail=False, methods=['POST'], permission_classes=[IsAuth])
    def scan(self, request):
        barcode = request.data.get("barcode")
        if not barcode:
            return Response({"error": "barcode required"}, status=400)
        try:
            asset = Asset.objects.select_related("category", "location", "department").get(barcode=barcode)
        except Asset.DoesNotExist:
            return Response({"error": "Asset not found"}, status=404)
        serializer = AssetDetailSerializer(asset)
        return Response({"found": True, "asset": serializer.data})

    @action(detail=False, methods=['POST'], permission_classes=[IsManagerOrAbove])
    def generate_barcodes(self, request):
        asset_ids = request.data.get("asset_ids", [])
        if not asset_ids or not isinstance(asset_ids, list):
            return Response({"error": "asset_ids list is required"}, status=400)
        task = generate_barcode_excel_task.delay(asset_ids)
        return Response(
            {"task_id": task.id, "message": "Barcode generation started"},
            status=status.HTTP_202_ACCEPTED,
        )


from rest_framework.decorators import api_view, permission_classes


@api_view(["GET"])
@permission_classes([IsAuth])
def approve_email(request, token):
    try:
        asset = Asset.objects.get(approval_token=token)
    except Asset.DoesNotExist:
        return error_response("Invalid token", http_status=status.HTTP_404_NOT_FOUND)
    if asset.approval_status != "PENDING":
        return error_response("Asset is not pending approval", http_status=status.HTTP_400_BAD_REQUEST)
    asset.approval_status = "APPROVED"
    asset.status = "ACTIVE"
    asset.save()
    return success_response(message="Asset approved successfully")


@api_view(["GET"])
@permission_classes([IsAuth])
def reject_email(request, token):
    try:
        asset = Asset.objects.get(approval_token=token)
    except Asset.DoesNotExist:
        return error_response("Invalid token", http_status=status.HTTP_404_NOT_FOUND)
    if asset.approval_status != "PENDING":
        return error_response("Asset is not pending approval", http_status=status.HTTP_400_BAD_REQUEST)
    asset.approval_status = "REJECTED"
    asset.status = "BLOCKED"
    asset.save()
    return success_response(message="Asset rejected")
