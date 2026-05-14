import io
import os
from datetime import date, timedelta
from decimal import Decimal

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
import numpy as np

from django.conf import settings
from django.db.models import Count, Sum
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether,
)
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from accounts.models import User
from assets.models import Asset, Category, AssetService, ServiceType
from repairs.models import RepairTicket
from procurement.models import ProcurementRequest
from .models import Report

# ─────────────────────────────────────────────
#  DESIGN TOKENS
# ─────────────────────────────────────────────
DARK_BG    = '#0a0f1e'
PANEL_BG   = '#0d1526'
CARD_BG    = '#111d35'
BORDER_COL = '#1e3a5f'
ACCENT     = '#00d4ff'
ACCENT2    = '#7c3aed'
GREEN      = '#00f5a0'
YELLOW     = '#ffd60a'
RED        = '#ff4d6d'
ORANGE     = '#ff6b35'
TEXT       = '#e8f4fd'
MUTED      = '#5b7fa6'
WHITE      = '#ffffff'

# ReportLab color objects
RL_DARK    = colors.HexColor(DARK_BG)
RL_PANEL   = colors.HexColor(PANEL_BG)
RL_CARD    = colors.HexColor(CARD_BG)
RL_BORDER  = colors.HexColor(BORDER_COL)
RL_ACCENT  = colors.HexColor(ACCENT)
RL_ACCENT2 = colors.HexColor(ACCENT2)
RL_GREEN   = colors.HexColor(GREEN)
RL_YELLOW  = colors.HexColor(YELLOW)
RL_RED     = colors.HexColor(RED)
RL_TEXT    = colors.HexColor(TEXT)
RL_MUTED   = colors.HexColor(MUTED)
RL_WHITE   = colors.HexColor(WHITE)

# Matplotlib global style
plt.rcParams.update({
    'figure.facecolor':  DARK_BG,
    'axes.facecolor':    PANEL_BG,
    'axes.edgecolor':    BORDER_COL,
    'axes.labelcolor':   MUTED,
    'text.color':        TEXT,
    'xtick.color':       MUTED,
    'ytick.color':       MUTED,
    'figure.dpi':        150,
    'font.family':       'DejaVu Sans',
    'grid.color':        BORDER_COL,
    'grid.linewidth':    0.5,
})

# Excel hex (no #)
_E_DARK   = '0A0F1E'
_E_PANEL  = '0D1526'
_E_CARD   = '111D35'
_E_ACCENT = '00D4FF'
_E_ACCENT2= '7C3AED'
_E_GREEN  = '00F5A0'
_E_YELLOW = 'FFD60A'
_E_RED    = 'FF4D6D'
_E_TEXT   = 'E8F4FD'
_E_MUTED  = '5B7FA6'
_E_BORDER = '1E3A5F'


# ─────────────────────────────────────────────
#  DATA COLLECTION
# ─────────────────────────────────────────────
def _scope(user):
    if user and getattr(user, 'role', None) == "SUPER_ADMIN":
        return {}
    if user and user.department:
        return {'department': user.department}
    return {'pk__in': []}


def collect_kpi_metrics(user=None):
    scope   = _scope(user)
    a_scope = {} if user and getattr(user, 'role', None) == "SUPER_ADMIN" else scope

    asset_qs     = Asset.objects.filter(**a_scope)
    total        = asset_qs.count()
    active       = asset_qs.filter(status='ACTIVE').count()
    under_repair = asset_qs.filter(status='REPAIR').count()
    missing      = asset_qs.filter(status='MISSING').count()
    retired      = asset_qs.filter(status='RETIRED').count()
    blocked      = asset_qs.filter(status='BLOCKED').count()

    pending_approvals = asset_qs.filter(approval_status='PENDING').count()

    proc_qs       = ProcurementRequest.objects.filter(**scope)
    total_procurements    = proc_qs.count()
    pending_procurements  = proc_qs.filter(approval_status='PENDING').count()
    approved_procurements = proc_qs.filter(approval_status='APPROVED').count()
    rejected_procurements = proc_qs.filter(approval_status='REJECTED').count()

    if not scope:
        repair_qs  = RepairTicket.objects.all()
    elif 'department' in scope:
        repair_qs  = RepairTicket.objects.filter(asset__department=scope['department'])
    else:
        repair_qs  = RepairTicket.objects.none()
    total_repairs       = repair_qs.count()
    open_repairs        = repair_qs.filter(status='OPEN').count()
    in_progress_repairs = repair_qs.filter(status='IN_PROGRESS').count()
    completed_repairs   = repair_qs.filter(status='COMPLETED').count()
    total_repair_cost   = repair_qs.aggregate(t=Sum('repair_cost'))['t'] or 0

    categories = Category.objects.annotate(asset_count=Count('assets')).values('name', 'asset_count')
    if a_scope:
        categories = Category.objects.filter(
            assets__department=scope.get('department')
        ).annotate(asset_count=Count('assets')).values('name', 'asset_count')

    last_week          = timezone.now() - timedelta(days=7)
    repairs_last_week  = repair_qs.filter(start_date__gte=last_week).count()
    proc_last_week     = proc_qs.filter(created_at__gte=last_week).count()

    today = timezone.now().date()
    svc_qs = AssetService.objects
    if a_scope:
        svc_qs = AssetService.objects.filter(asset__department=scope.get('department'))
    total_services      = svc_qs.filter(end_date__gte=today).count()
    expiring_30         = svc_qs.filter(
        end_date__gte=today,
        end_date__lte=today + timedelta(days=30),
    ).count()
    expiring_7          = svc_qs.filter(
        end_date__gte=today,
        end_date__lte=today + timedelta(days=7),
    ).count()
    expired_services    = svc_qs.filter(end_date__lt=today).count()
    expiring_by_type    = (
        svc_qs.filter(end_date__gte=today,
            end_date__lte=today + timedelta(days=30))
        .values('service_type__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    expiring_details    = [
        {**ed, 'end_date': str(ed['end_date'])}
        for ed in svc_qs.filter(
            end_date__gte=today,
            end_date__lte=today + timedelta(days=30),
        ).select_related('asset', 'service_type').values(
            'asset__asset_code', 'asset__asset_name',
            'service_type__name', 'end_date',
        )[:20]
    ]

    utilisation_rate = round((active / total * 100), 1) if total else 0
    repair_rate      = round((under_repair / total * 100), 1) if total else 0
    compliance_score = round(
        ((total_services - expired_services) / total_services * 100), 1
    ) if total_services else 100

    return {
        'total_assets':              total,
        'active_assets':             active,
        'under_repair':              under_repair,
        'missing_assets':            missing,
        'retired_assets':            retired,
        'blocked_assets':            blocked,
        'pending_approvals':         pending_approvals,
        'total_procurements':        total_procurements,
        'pending_procurements':      pending_procurements,
        'approved_procurements':     approved_procurements,
        'rejected_procurements':     rejected_procurements,
        'total_repairs':             total_repairs,
        'open_repairs':              open_repairs,
        'in_progress_repairs':       in_progress_repairs,
        'completed_repairs':         completed_repairs,
        'total_repair_cost':         float(total_repair_cost),
        'categories':                list(categories),
        'repairs_last_week':         repairs_last_week,
        'proc_last_week':            proc_last_week,
        'utilisation_rate':          utilisation_rate,
        'repair_rate':               repair_rate,
        'total_services':            total_services,
        'expiring_30':               expiring_30,
        'expiring_7':                expiring_7,
        'expired_services':          expired_services,
        'expiring_by_type':          list(expiring_by_type),
        'expiring_details':          expiring_details,
        'compliance_score':          compliance_score,
    }


# ─────────────────────────────────────────────
#  CHART GENERATION  (matplotlib → PNG bytes)
# ─────────────────────────────────────────────
def _save_fig(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', facecolor=DARK_BG, dpi=150)
    buf.seek(0)
    data = buf.getvalue()
    plt.close(fig)
    return data


def _glow_bar(ax, x, height, width=0.55, color=ACCENT, alpha_core=0.9):
    """Draw a bar with a subtle glow halo."""
    # halo layers
    for a, w in [(0.08, width + 0.22), (0.14, width + 0.12), (0.22, width + 0.04)]:
        ax.bar(x, height, width=w, color=color, alpha=a, zorder=1)
    ax.bar(x, height, width=width, color=color, alpha=alpha_core, zorder=2)


def chart_asset_status(metrics: dict) -> bytes:
    status_map = {
        'Active':      (metrics['active_assets'],   GREEN),
        'In Repair':   (metrics['under_repair'],     YELLOW),
        'Missing':     (metrics['missing_assets'],   RED),
        'Retired':     (metrics['retired_assets'],   MUTED),
        'Blocked':     (metrics['blocked_assets'],   ORANGE),
    }
    labels = [k for k, (v, _) in status_map.items() if v > 0]
    vals   = [v for v, _ in status_map.values() if v > 0]
    clrs   = [c for v, c in status_map.values() if v > 0]

    fig, ax = plt.subplots(figsize=(6, 5))
    fig.patch.set_facecolor(DARK_BG)

    wedge_props = dict(width=0.55, edgecolor=DARK_BG, linewidth=2)
    wedges, _, autotexts = ax.pie(
        vals, colors=clrs, autopct='%1.1f%%',
        pctdistance=0.75, startangle=110,
        wedgeprops=wedge_props,
        textprops={'color': TEXT, 'fontsize': 9, 'fontweight': 'bold'},
    )
    # centre label
    ax.text(0, 0, f'{sum(vals)}\nTotal', ha='center', va='center',
            color=WHITE, fontsize=13, fontweight='bold', linespacing=1.5)

    ax.legend(
        [f'{l}  ({v})' for l, v in zip(labels, vals)],
        loc='lower center', bbox_to_anchor=(0.5, -0.14),
        ncol=3, fontsize=8, facecolor=CARD_BG,
        edgecolor=BORDER_COL, labelcolor=TEXT, framealpha=0.9,
    )
    ax.set_title('Asset Status Overview', color=ACCENT, fontsize=14,
                 fontweight='bold', pad=18)
    return _save_fig(fig)


def chart_category_dist(metrics: dict) -> bytes:
    cats = sorted(metrics['categories'], key=lambda c: c['asset_count'], reverse=True)
    if not cats:
        cats = [{'name': 'No Data', 'asset_count': 0}]
    names = [c['name'] for c in cats]
    vals  = [c['asset_count'] for c in cats]
    n     = len(names)
    clrs  = [plt.cm.cool(i / max(n - 1, 1)) for i in range(n)]

    fig, ax = plt.subplots(figsize=(7, max(3, n * 0.55 + 1)))
    for i, (name, val, clr) in enumerate(zip(names, vals, clrs)):
        # background track
        ax.barh(i, max(vals) * 1.1 if max(vals) else 1, color=PANEL_BG,
                height=0.55, left=0, zorder=1)
        # filled bar
        ax.barh(i, val, color=clr, height=0.55, zorder=2, alpha=0.88)
        ax.text(val + max(vals) * 0.02 if max(vals) else 0.5, i,
                str(val), va='center', fontsize=9,
                color=TEXT, fontweight='bold', zorder=3)

    ax.set_yticks(range(n))
    ax.set_yticklabels(names, fontsize=9)
    ax.set_xlabel('Asset Count', fontsize=9, color=MUTED)
    ax.set_xlim(0, max(vals) * 1.2 if max(vals) else 2)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='x', alpha=0.3, zorder=0)
    ax.set_title('Assets by Category', color=ACCENT, fontsize=14,
                 fontweight='bold', pad=12)
    fig.tight_layout()
    return _save_fig(fig)


def chart_repair_status(metrics: dict) -> bytes:
    labels = ['Open', 'In Progress', 'Completed']
    vals   = [metrics['open_repairs'], metrics['in_progress_repairs'],
              metrics['completed_repairs']]
    clrs   = [RED, YELLOW, GREEN]

    fig, ax = plt.subplots(figsize=(6, 3.5))
    x = np.arange(len(labels))
    for xi, (v, c) in enumerate(zip(vals, clrs)):
        _glow_bar(ax, xi, v, color=c)
        ax.text(xi, v + max(vals) * 0.04 if max(vals) else 0.5, str(v),
                ha='center', fontsize=12, color=c, fontweight='bold', zorder=5)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=10)
    ax.set_ylabel('Ticket Count', fontsize=9, color=MUTED)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', alpha=0.25, zorder=0)
    ax.set_ylim(0, max(vals) * 1.3 if max(vals) else 5)
    ax.set_title('Repair Tickets by Status', color=ACCENT, fontsize=14,
                 fontweight='bold', pad=12)
    fig.tight_layout()
    return _save_fig(fig)


def chart_procurement_overview(metrics: dict) -> bytes:
    labels = ['Pending', 'Approved', 'Rejected']
    vals   = [metrics['pending_procurements'],
              metrics['approved_procurements'],
              metrics['rejected_procurements']]
    clrs   = [YELLOW, GREEN, RED]

    fig, ax = plt.subplots(figsize=(5, 3.5))
    wedge_props = dict(edgecolor=DARK_BG, linewidth=2)
    non_zero = [(l, v, c) for l, v, c in zip(labels, vals, clrs) if v > 0]
    if not non_zero:
        non_zero = [('No Data', 1, MUTED)]
    ls, vs, cs = zip(*non_zero)
    wedges, _, autotexts = ax.pie(
        vs, colors=cs, autopct='%1.0f%%', startangle=90,
        wedgeprops=wedge_props,
        textprops={'color': TEXT, 'fontsize': 9, 'fontweight': 'bold'},
    )
    ax.legend(
        [f'{l} ({v})' for l, v in zip(ls, vs)],
        loc='lower center', bbox_to_anchor=(0.5, -0.12),
        ncol=3, fontsize=8, facecolor=CARD_BG,
        edgecolor=BORDER_COL, labelcolor=TEXT,
    )
    ax.set_title('Procurement Status', color=ACCENT, fontsize=14,
                 fontweight='bold', pad=12)
    return _save_fig(fig)


def chart_compliance(metrics: dict) -> bytes:
    labels = ['Active', 'Expiring (7d)', 'Expiring (30d)', 'Expired']
    vals   = [
        max(0, metrics['total_services'] - metrics['expiring_30'] - metrics['expired_services']),
        metrics['expiring_7'],
        max(0, metrics['expiring_30'] - metrics['expiring_7']),
        metrics['expired_services'],
    ]
    clrs   = [GREEN, YELLOW, ACCENT, RED]

    fig, ax = plt.subplots(figsize=(5, 3.5))
    wedge_props = dict(edgecolor=DARK_BG, linewidth=2)
    non_zero = [(l, v, c) for l, v, c in zip(labels, vals, clrs) if v > 0]
    if not non_zero:
        non_zero = [('No Services', 1, MUTED)]
    ls, vs, cs = zip(*non_zero)
    wedges, _, autotexts = ax.pie(
        vs, colors=cs, autopct='%1.0f%%', startangle=90,
        wedgeprops=wedge_props,
        textprops={'color': TEXT, 'fontsize': 9, 'fontweight': 'bold'},
    )
    ax.legend(
        [f'{l} ({v})' for l, v in zip(ls, vs)],
        loc='lower center', bbox_to_anchor=(0.5, -0.12),
        ncol=2, fontsize=8, facecolor=CARD_BG,
        edgecolor=BORDER_COL, labelcolor=TEXT,
    )
    ax.set_title('Service Compliance Overview', color=ACCENT, fontsize=14,
                 fontweight='bold', pad=12)
    return _save_fig(fig)


def generate_charts(metrics: dict) -> dict:
    return {
        'asset_status':   chart_asset_status(metrics),
        'category_dist':  chart_category_dist(metrics),
        'repair_status':  chart_repair_status(metrics),
        'procurement':    chart_procurement_overview(metrics),
        'compliance':     chart_compliance(metrics),
    }


# ─────────────────────────────────────────────
#  PDF – CANVAS BACKGROUND HELPERS
# ─────────────────────────────────────────────
PAGE_W, PAGE_H = A4   # 595.28 × 841.89 pt

def _draw_page_background(canv, doc):
    """Full-page dark background + subtle decorative lines."""
    canv.saveState()
    canv.setFillColor(RL_DARK)
    canv.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

    # top accent bar
    canv.setFillColor(RL_ACCENT)
    canv.rect(0, PAGE_H - 3, PAGE_W, 3, fill=1, stroke=0)

    # bottom accent bar
    canv.setFillColor(RL_ACCENT2)
    canv.rect(0, 0, PAGE_W, 2, fill=1, stroke=0)

    # faint diagonal grid lines (decorative)
    canv.setStrokeColor(colors.HexColor(BORDER_COL))
    canv.setLineWidth(0.3)
    canv.setDash([2, 8])
    for x in range(-100, int(PAGE_W) + 100, 60):
        canv.line(x, 0, x + 120, PAGE_H)
    canv.setDash()
    canv.restoreState()


def _draw_cover(canv, title: str, subtitle: str, meta: dict = None):
    """Draw a striking cover page directly on the canvas."""
    canv.saveState()
    meta = meta or {}

    # background
    canv.setFillColor(RL_DARK)
    canv.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

    # top gradient-like bars
    for i, (clr, h) in enumerate([(ACCENT, 5), (ACCENT2, 3), (BORDER_COL, 1)]):
        canv.setFillColor(colors.HexColor(clr))
        canv.rect(0, PAGE_H - sum(x[1] for x in [(ACCENT, 5), (ACCENT2, 3), (BORDER_COL, 1)][:i+1]),
                  PAGE_W, h, fill=1, stroke=0)

    # big glowing circle background
    canv.setFillColor(colors.HexColor('#05162e'))
    canv.circle(PAGE_W / 2, PAGE_H / 2 + 60, 190, fill=1, stroke=0)
    canv.setStrokeColor(colors.HexColor(ACCENT))
    canv.setLineWidth(1.2)
    canv.circle(PAGE_W / 2, PAGE_H / 2 + 60, 190, fill=0, stroke=1)
    canv.setStrokeColor(colors.HexColor(ACCENT2))
    canv.setLineWidth(0.5)
    canv.circle(PAGE_W / 2, PAGE_H / 2 + 60, 210, fill=0, stroke=1)

    # organisation / logo placeholder
    canv.setFillColor(RL_ACCENT)
    canv.setFont('Helvetica-Bold', 11)
    canv.drawCentredString(PAGE_W / 2, PAGE_H - 40, '◈  ASSET MANAGEMENT SYSTEM')

    # divider line
    canv.setStrokeColor(RL_ACCENT)
    canv.setLineWidth(0.8)
    canv.line(PAGE_W * 0.2, PAGE_H - 50, PAGE_W * 0.8, PAGE_H - 50)

    # main title
    canv.setFont('Helvetica-Bold', 32)
    canv.setFillColor(RL_WHITE)
    # word-wrap title manually
    lines = _wrap_text(title, 26)
    y = PAGE_H / 2 + 120
    for line in lines:
        canv.drawCentredString(PAGE_W / 2, y, line)
        y -= 40

    # accent underline
    canv.setStrokeColor(RL_ACCENT)
    canv.setLineWidth(2)
    canv.line(PAGE_W * 0.3, y + 10, PAGE_W * 0.7, y + 10)

    # subtitle / date
    canv.setFont('Helvetica', 12)
    canv.setFillColor(RL_MUTED)
    canv.drawCentredString(PAGE_W / 2, y - 12, subtitle)

    # metadata block
    meta_lines = []
    if meta.get('report_id'):
        meta_lines.append(f"Report ID: {meta['report_id']}")
    if meta.get('department'):
        meta_lines.append(f"Department: {meta['department']}")
    if meta.get('version'):
        meta_lines.append(f"Version: {meta['version']}")
    if meta.get('generated_by'):
        meta_lines.append(f"Generated By: {meta['generated_by']}")
    canv.setFont('Helvetica', 9)
    canv.setFillColor(RL_MUTED)
    my = y - 36
    for ml in meta_lines:
        canv.drawCentredString(PAGE_W / 2, my, ml)
        my -= 14

    # KPI hint box
    box_y = PAGE_H / 2 - 100
    canv.setFillColor(colors.HexColor(CARD_BG))
    canv.roundRect(PAGE_W * 0.15, box_y, PAGE_W * 0.70, 60, 8, fill=1, stroke=0)
    canv.setStrokeColor(RL_BORDER)
    canv.setLineWidth(0.8)
    canv.roundRect(PAGE_W * 0.15, box_y, PAGE_W * 0.70, 60, 8, fill=0, stroke=1)
    canv.setFont('Helvetica', 9)
    canv.setFillColor(RL_MUTED)
    canv.drawCentredString(PAGE_W / 2, box_y + 42,
                           'CONFIDENTIAL  |  INTERNAL USE ONLY')
    canv.setFont('Helvetica-Bold', 10)
    canv.setFillColor(RL_TEXT)
    canv.drawCentredString(PAGE_W / 2, box_y + 22,
                           f'Report Date: {date.today().strftime("%A, %B %d, %Y")}')

    # bottom bar
    canv.setFillColor(RL_ACCENT2)
    canv.rect(0, 0, PAGE_W, 4, fill=1, stroke=0)
    canv.setFillColor(RL_MUTED)
    canv.setFont('Helvetica', 8)
    canv.drawCentredString(PAGE_W / 2, 10, 'Generated automatically · For authorised personnel only')

    canv.restoreState()


def _wrap_text(text: str, max_chars: int) -> list:
    words = text.split()
    lines, current = [], ''
    for w in words:
        if len(current) + len(w) + 1 <= max_chars:
            current = (current + ' ' + w).strip()
        else:
            if current:
                lines.append(current)
            current = w
    if current:
        lines.append(current)
    return lines or [text]


# ─────────────────────────────────────────────
#  PDF – STYLES
# ─────────────────────────────────────────────
def _build_styles():
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    return {
        'section_title': ps('SectionTitle',
            fontSize=16, textColor=RL_ACCENT,
            spaceAfter=8, spaceBefore=20,
            fontName='Helvetica-Bold'),

        'sub_title': ps('SubTitle',
            fontSize=11, textColor=RL_MUTED,
            spaceAfter=14, fontName='Helvetica'),

        'body': ps('Body',
            fontSize=9, textColor=RL_TEXT,
            spaceAfter=5, fontName='Helvetica',
            leading=14),

        'insight': ps('Insight',
            fontSize=9.5, textColor=RL_TEXT,
            spaceAfter=6, leftIndent=14,
            fontName='Helvetica', leading=15),

        'caption': ps('Caption',
            fontSize=8, textColor=RL_MUTED,
            spaceAfter=4, alignment=TA_CENTER,
            fontName='Helvetica-Oblique'),

        'footer_text': ps('FooterText',
            fontSize=7, textColor=RL_MUTED,
            alignment=TA_CENTER, fontName='Helvetica'),

        'kpi_label': ps('KpiLabel',
            fontSize=8, textColor=RL_MUTED,
            fontName='Helvetica', spaceAfter=2),

        'kpi_value': ps('KpiValue',
            fontSize=18, textColor=RL_ACCENT,
            fontName='Helvetica-Bold', spaceAfter=4),
    }


# ─────────────────────────────────────────────
#  PDF – KPI CARDS TABLE
# ─────────────────────────────────────────────
def _kpi_cards_table(metrics: dict, styles: dict):
    """Return a Table that renders as a 4-col grid of KPI cards."""
    CARD_W = 121
    kpis = [
        ('TOTAL ASSETS',      str(metrics['total_assets']),        RL_ACCENT),
        ('ACTIVE',            str(metrics['active_assets']),        RL_GREEN),
        ('UNDER REPAIR',      str(metrics['under_repair']),         RL_YELLOW),
        ('MISSING',           str(metrics['missing_assets']),       RL_RED),
        ('UTILISATION',       f"{metrics['utilisation_rate']}%",    RL_ACCENT),
        ('REPAIR RATE',       f"{metrics['repair_rate']}%",         RL_YELLOW),
        ('PENDING APPROVALS', str(metrics['pending_approvals']),    RL_MUTED),
        ('REPAIR COST',       f"${metrics['total_repair_cost']:,.0f}", RL_ACCENT2),
    ]
    cells = []
    row = []
    for i, (label, value, color) in enumerate(kpis):
        cell_content = Table(
            [[Paragraph(label, styles['kpi_label'])],
             [Paragraph(f'<font color="{color.hexval()}">{value}</font>',
                        styles['kpi_value'])]],
            colWidths=[CARD_W - 12],
        )
        cell_content.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        row.append(cell_content)
        if len(row) == 4:
            cells.append(row)
            row = []
    if row:
        while len(row) < 4:
            row.append(Paragraph('', styles['body']))
        cells.append(row)

    t = Table(cells, colWidths=[CARD_W] * 4, rowHeights=[64] * len(cells))
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, -1), RL_CARD),
        ('BOX',          (0, 0), (-1, -1), 1.2, RL_BORDER),
        ('INNERGRID',    (0, 0), (-1, -1), 0.6, RL_BORDER),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING',   (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 8),
        # accent top border on each cell via row background
        ('LINEABOVE',    (0, 0), (-1, 0), 2, RL_ACCENT),
    ]))
    return t


# ─────────────────────────────────────────────
#  PDF – SECTION DIVIDER
# ─────────────────────────────────────────────
def _section_divider(label: str, styles: dict):
    return KeepTogether([
        HRFlowable(width='100%', color=RL_BORDER, thickness=0.5,
                   spaceAfter=0, spaceBefore=4),
        Paragraph(f'<font color="{ACCENT}">◆</font>  {label}',
                  styles['section_title']),
    ])


# ─────────────────────────────────────────────
#  PDF – ASSET DETAIL TABLE
# ─────────────────────────────────────────────
def _asset_detail_table(styles: dict, user=None):
    from assets.models import Asset
    headers = ['Code', 'Asset Name', 'Status', 'Category', 'Location']
    col_w   = [65, 130, 65, 90, 90]
    data    = [headers]

    status_colors = {
        'ACTIVE':   GREEN,  'REPAIR': YELLOW,
        'MISSING':  RED,    'RETIRED': MUTED,
        'BLOCKED':  ORANGE,
    }

    a_qs = Asset.objects.select_related('category', 'location')
    if user and getattr(user, 'role', None) != "SUPER_ADMIN" and user.department:
        a_qs = a_qs.filter(department=user.department)
    assets = a_qs[:60]
    for a in assets:
        sc = status_colors.get(a.status, TEXT)
        data.append([
            a.asset_code,
            a.asset_name,
            Paragraph(f'<font color="{sc}">{a.status}</font>', styles['body']),
            a.category.name if a.category else '—',
            a.location.name if a.location else '—',
        ])

    t = Table(data, colWidths=col_w, repeatRows=1)
    ts = [
        # Header row
        ('BACKGROUND',    (0, 0), (-1, 0),  RL_PANEL),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  RL_ACCENT),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.5, RL_ACCENT),
        # Data rows
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('TEXTCOLOR',     (0, 1), (-1, -1), RL_TEXT),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [RL_DARK, RL_CARD]),
        ('GRID',          (0, 0), (-1, -1), 0.4, RL_BORDER),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]
    t.setStyle(TableStyle(ts))
    return t


# ─────────────────────────────────────────────
#  PDF – MAIN GENERATOR
# ─────────────────────────────────────────────
def generate_pdf(metrics: dict, charts: dict,
                 title: str = "Weekly Executive Report",
                 user=None) -> bytes:

    buf = io.BytesIO()

    # ── Page template callbacks ──────────────────────────────────────────
    def on_page(canv, doc):
        _draw_page_background(canv, doc)
        # page number footer
        canv.saveState()
        canv.setFillColor(RL_MUTED)
        canv.setFont('Helvetica', 7)
        page_num = doc.page
        canv.drawRightString(PAGE_W - 15*mm, 8*mm, f'Page {page_num}')
        canv.drawString(15*mm, 8*mm, 'CONFIDENTIAL  |  INTERNAL')
        canv.restoreState()

    # ── Document setup ───────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20*mm, bottomMargin=18*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    S = _build_styles()
    elements = []

    # ── COVER PAGE (raw canvas, then force page break) ───────────────────
    # We inject the cover via a canvas action inside the first page template.
    # Easiest with a temporary small doctemplate draw, then we inject via
    # onFirstPage / onLaterPages callbacks.

    def on_first_page(canv, doc):
        meta = {
            'report_id': f"WR-{date.today().strftime('%Y%m%d')}",
            'department': user.department.department_name if user and user.department else 'Organization',
            'version': '1.0',
            'generated_by': user.username if user else 'System',
        }
        _draw_cover(canv, title,
                    f"Generated: {date.today().strftime('%B %d, %Y')}",
                    meta=meta)

    def on_later_pages(canv, doc):
        on_page(canv, doc)

    # Cover page handled by on_first_page callback
    elements.append(PageBreak())

    # ── PAGE 2: EXECUTIVE SUMMARY ────────────────────────────────────────
    elements.append(_section_divider('Executive Summary', S))
    elements.append(Spacer(1, 6))

    summary_lines = [
        f"This report covers the full asset, repair, and procurement status as of "
        f"<b>{date.today().strftime('%B %d, %Y')}</b>.",
        f"The fleet currently comprises <b>{metrics['total_assets']}</b> assets with "
        f"a utilisation rate of <b>{metrics['utilisation_rate']}%</b>.",
        f"Repair activity stands at <b>{metrics['repair_rate']}%</b> of total inventory.",
    ]
    for line in summary_lines:
        elements.append(Paragraph(line, S['body']))

    elements.append(Spacer(1, 14))
    elements.append(_section_divider('Key Performance Indicators', S))
    elements.append(Spacer(1, 8))
    elements.append(_kpi_cards_table(metrics, S))

    # ── KEY INSIGHTS ─────────────────────────────────────────────────────
    elements.append(Spacer(1, 12))
    elements.append(_section_divider('Key Insights', S))

    insights = []
    if metrics['repairs_last_week']:
        insights.append(
            f"<font color='{YELLOW}'>▲</font>  "
            f"<b>{metrics['repairs_last_week']}</b> repair tickets opened in the last 7 days.")
    if metrics['proc_last_week']:
        insights.append(
            f"<font color='{ACCENT}'>▲</font>  "
            f"<b>{metrics['proc_last_week']}</b> new procurement requests this week.")
    if metrics['missing_assets']:
        insights.append(
            f"<font color='{RED}'>⚠</font>  "
            f"<b>{metrics['missing_assets']}</b> assets are currently marked as <b>MISSING</b> — "
            f"immediate review required.")
    if metrics['pending_approvals']:
        insights.append(
            f"<font color='{YELLOW}'>⏳</font>  "
            f"<b>{metrics['pending_approvals']}</b> asset approvals are still pending.")
    if metrics['pending_procurements']:
        insights.append(
            f"<font color='{YELLOW}'>⏳</font>  "
            f"<b>{metrics['pending_procurements']}</b> procurement requests are awaiting approval.")
    if not insights:
        insights.append(
            f"<font color='{GREEN}'>✔</font>  No critical issues detected this week.")

    for ins in insights:
        elements.append(Paragraph(ins, S['insight']))

    elements.append(PageBreak())

    # ── PAGE 3: CHARTS ───────────────────────────────────────────────────
    elements.append(_section_divider('Visual Analytics', S))
    elements.append(Spacer(1, 6))

    chart_specs = [
        ('asset_status',  'Figure 1 — Asset Status Distribution'),
        ('category_dist', 'Figure 2 — Assets by Category'),
        ('repair_status', 'Figure 3 — Repair Ticket Status'),
        ('procurement',   'Figure 4 — Procurement Approval Status'),
    ]

    # 2×2 layout: pair charts side by side
    pairs = [chart_specs[i:i+2] for i in range(0, len(chart_specs), 2)]
    for pair in pairs:
        row_imgs = []
        for key, caption in pair:
            if key in charts:
                img = Image(io.BytesIO(charts[key]), width=240, height=160)
                cap = Paragraph(caption, S['caption'])
                cell = Table([[img], [cap]], colWidths=[244])
                cell.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), RL_PANEL),
                    ('BOX',        (0, 0), (-1, -1), 0.6, RL_BORDER),
                    ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                row_imgs.append(cell)
        if len(row_imgs) == 2:
            row_t = Table([row_imgs], colWidths=[252, 252])
            row_t.setStyle(TableStyle([
                ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
                ('LEFTPADDING',  (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(row_t)
        elif row_imgs:
            elements.append(row_imgs[0])
        elements.append(Spacer(1, 10))

    elements.append(PageBreak())

    # ── COMPLIANCE & SERVICE EXPIRY PAGE ─────────────────────────────────
    if metrics.get('total_services', 0) > 0:
        elements.append(_section_divider('Compliance & Service Expiry', S))
        elements.append(Spacer(1, 8))

        comp_kpis = [
            ['Active Services', str(metrics['total_services']),
             'Compliance Score', f"{metrics['compliance_score']}%"],
            ['Expiring (30 days)', str(metrics['expiring_30']),
             'Expired', str(metrics['expired_services'])],
        ]
        comp_t = Table(comp_kpis, colWidths=[120, 70, 120, 70])
        comp_t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(TEXT)),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(CARD_BG)),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTNAME', (3, 0), (3, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, -1), 12),
            ('FONTSIZE', (3, 0), (3, -1), 12),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor(ACCENT)),
            ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor(ACCENT)),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(ACCENT)),
        ]))
        elements.append(comp_t)
        elements.append(Spacer(1, 12))

        if 'compliance' in charts:
            img = Image(io.BytesIO(charts['compliance']), width=460, height=260)
            elements.append(img)
            elements.append(Spacer(1, 12))

        exp_details = metrics.get('expiring_details', [])
        expiring_count = metrics.get('expiring_30', 0)
        expiring_by_type = metrics.get('expiring_by_type', [])

        if expiring_by_type:
            elements.append(_section_divider('Services Expiring Soon — Breakdown by Type', S))
            type_headers = ['Service Type', 'Count']
            type_data = [type_headers]
            for et in expiring_by_type:
                type_data.append([et['service_type__name'], str(et['count'])])
            type_data.append(['Total', str(expiring_count)])
            type_t = Table(type_data, colWidths=[250, 100])
            type_t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(RED)),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor(TEXT)),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(CARD_BG)),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a1a3a')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(type_t)
            elements.append(Spacer(1, 10))

        elements.append(_section_divider('Services Expiring in Next 30 Days', S))
        if exp_details:
            exp_headers = ['Asset Code', 'Asset', 'Service', 'Expiry', 'Days Left']
            exp_data = [exp_headers]
            today = date.today()
            for ed in exp_details[:15]:
                exp_d = date.fromisoformat(ed['end_date']) if isinstance(ed['end_date'], str) else ed['end_date']
                days_left = (exp_d - today).days if exp_d else 0
                risk = 'CRITICAL' if days_left <= 7 else 'HIGH' if days_left <= 15 else 'WARNING'
                exp_data.append([
                    ed['asset__asset_code'], ed['asset__asset_name'][:25],
                    ed['service_type__name'], str(exp_d), f"{days_left}d ({risk})",
                ])
            exp_t = Table(exp_data, colWidths=[75, 120, 80, 70, 80])
            exp_t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(RED)),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor(TEXT)),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(CARD_BG)),
                ('ALIGN', (3, 0), (4, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
            ]))
            elements.append(exp_t)
        else:
            elements.append(Paragraph(
                f'<para alignment="center"><font color="{MUTED}" size="10">No services are expiring in the next 30 days. All service agreements are current.</font></para>',
                S['body'],
            ))

        elements.append(PageBreak())

    # ── PAGE 4+: ASSET DETAIL TABLE ─────────────────────────────────────
    elements.append(_section_divider('Detailed Asset Register (Top 60)', S))
    elements.append(Spacer(1, 8))
    elements.append(_asset_detail_table(S, user=user))

    # ── REPAIRS SUMMARY TABLE ─────────────────────────────────────────────
    elements.append(Spacer(1, 16))
    elements.append(_section_divider('Repair Ticket Summary', S))
    elements.append(Spacer(1, 8))

    repair_headers = ['Ticket ID', 'Asset Code', 'Status', 'Cost', 'Started', 'Completed']
    r_col_w = [55, 80, 70, 70, 75, 75]
    repair_data = [repair_headers]
    repair_status_colors = {
        'OPEN': RED, 'IN_PROGRESS': YELLOW, 'COMPLETED': GREEN,
    }
    last_week = timezone.now() - timedelta(days=7)

    repairs = (
        RepairTicket.objects
        .select_related('asset')
        .filter(start_date__gte=last_week)
        .order_by('-start_date')[:40]
    )
    for r in repairs:
        sc = repair_status_colors.get(r.status, TEXT)
        repair_data.append([
            str(r.id),
            r.asset.asset_code if r.asset else '—',
            Paragraph(f'<font color="{sc}">{r.status}</font>',
                      S['body']),
            f'${float(r.repair_cost):,.2f}',
            r.start_date.strftime('%Y-%m-%d') if r.start_date else '—',
            r.completion_date.strftime('%Y-%m-%d') if r.completion_date else '—',
        ])

    rt = Table(repair_data, colWidths=r_col_w, repeatRows=1)
    rt.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  RL_PANEL),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  RL_ACCENT),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.5, RL_ACCENT),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('TEXTCOLOR',     (0, 1), (-1, -1), RL_TEXT),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [RL_DARK, RL_CARD]),
        ('GRID',          (0, 0), (-1, -1), 0.4, RL_BORDER),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]))
    elements.append(rt)

    # ── BUILD ────────────────────────────────────────────────────────────
    doc.build(
        elements,
        onFirstPage=on_first_page,
        onLaterPages=on_later_pages,
    )
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  EXCEL – FUTURISTIC DARK THEME
# ─────────────────────────────────────────────
def _xfont(bold=False, size=10, color=_E_TEXT, italic=False):
    return Font(bold=bold, size=size,
                color=color, italic=italic,
                name='Calibri')

def _xfill(hex_color: str):
    return PatternFill(start_color=hex_color,
                       end_color=hex_color, fill_type='solid')

def _xborder(color=_E_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _xalign(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_cell(ws, row, col, value,
                 bg=_E_PANEL, fg=_E_ACCENT, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font    = _xfont(bold=True, size=size, color=fg)
    cell.fill    = _xfill(bg)
    cell.border  = _xborder()
    cell.alignment = _xalign('center')
    return cell

def _data_cell(ws, row, col, value, bg=None, fg=_E_TEXT, bold=False):
    fill_bg = bg if bg else (_E_DARK if row % 2 == 0 else _E_CARD)
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _xfont(bold=bold, color=fg)
    cell.fill      = _xfill(fill_bg)
    cell.border    = _xborder()
    cell.alignment = _xalign('left')
    return cell


def generate_excel(metrics: dict, title: str = "Weekly Report", user=None) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Summary Dashboard ───────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value=f"◈  {title.upper()}")
    c.font = _xfont(bold=True, size=16, color=_E_ACCENT)
    c.fill = _xfill(_E_DARK)
    c.alignment = _xalign('left')

    ws.merge_cells('A2:F2')
    c2 = ws.cell(row=2, column=1,
                 value=f"Generated: {date.today().strftime('%A, %B %d, %Y')}  |  CONFIDENTIAL")
    c2.font = _xfont(size=9, color=_E_MUTED, italic=True)
    c2.fill = _xfill(_E_DARK)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # Spacer row
    for col in range(1, 7):
        ws.cell(row=3, column=col).fill = _xfill(_E_DARK)

    # KPI section header
    ws.merge_cells('A4:F4')
    sh = ws.cell(row=4, column=1, value='KEY PERFORMANCE INDICATORS')
    sh.font = _xfont(bold=True, size=10, color=_E_ACCENT)
    sh.fill = _xfill(_E_PANEL)
    sh.alignment = _xalign('center')

    kpi_rows = [
        ('Total Assets',          metrics['total_assets'],         _E_ACCENT),
        ('Active Assets',         metrics['active_assets'],        _E_GREEN),
        ('Under Repair',          metrics['under_repair'],         _E_YELLOW),
        ('Missing Assets',        metrics['missing_assets'],       _E_RED),
        ('Retired Assets',        metrics['retired_assets'],       _E_MUTED),
        ('Blocked Assets',        metrics['blocked_assets'],       _E_MUTED),
        ('Utilisation Rate',      f"{metrics['utilisation_rate']}%", _E_ACCENT),
        ('Repair Rate',           f"{metrics['repair_rate']}%",    _E_YELLOW),
        ('Pending Approvals',     metrics['pending_approvals'],    _E_YELLOW),
        ('Total Procurements',    metrics['total_procurements'],   _E_TEXT),
        ('Pending Procurements',  metrics['pending_procurements'], _E_YELLOW),
        ('Approved Procurements', metrics['approved_procurements'],_E_GREEN),
        ('Rejected Procurements', metrics['rejected_procurements'],_E_RED),
        ('Total Repairs',         metrics['total_repairs'],        _E_TEXT),
        ('Open Repairs',          metrics['open_repairs'],         _E_RED),
        ('In-Progress Repairs',   metrics['in_progress_repairs'],  _E_YELLOW),
        ('Completed Repairs',     metrics['completed_repairs'],    _E_GREEN),
        ('Total Repair Cost',     f"${metrics['total_repair_cost']:,.2f}", _E_ACCENT),
    ]

    for i, (label, value, color) in enumerate(kpi_rows, start=5):
        bg = _E_DARK if i % 2 == 0 else _E_CARD
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _xfont(color=_E_TEXT)
        lc.fill = _xfill(bg)
        lc.border = _xborder()
        lc.alignment = _xalign('left')

        vc = ws.cell(row=i, column=2, value=value)
        vc.font = _xfont(bold=True, color=color)
        vc.fill = _xfill(bg)
        vc.border = _xborder()
        vc.alignment = _xalign('center')

        ws.row_dimensions[i].height = 20

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22

    # ── Sheet 2: Asset Register ───────────────────────────────────────────
    ws2 = wb.create_sheet("Asset Register")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:G1')
    title_c = ws2.cell(row=1, column=1, value='ASSET REGISTER')
    title_c.font = _xfont(bold=True, size=13, color=_E_ACCENT)
    title_c.fill = _xfill(_E_DARK)
    ws2.row_dimensions[1].height = 24

    hdrs = ['Asset Code', 'Asset Name', 'Brand', 'Status', 'Category', 'Location', 'Approval']
    for col, h in enumerate(hdrs, 1):
        _header_cell(ws2, 2, col, h)
    ws2.row_dimensions[2].height = 18

    STATUS_COLORS_XL = {
        'ACTIVE': _E_GREEN, 'REPAIR': _E_YELLOW,
        'MISSING': _E_RED,  'RETIRED': _E_MUTED,
        'BLOCKED': 'FF6B35',
    }
    xa_qs = Asset.objects.select_related('category', 'location')
    if user and getattr(user, 'role', None) != "SUPER_ADMIN" and user.department:
        xa_qs = xa_qs.filter(department=user.department)
    assets = xa_qs.all()
    for i, a in enumerate(assets, start=3):
        sc = STATUS_COLORS_XL.get(a.status, _E_TEXT)
        bg = _E_DARK if i % 2 == 0 else _E_CARD
        row_data = [
            (a.asset_code,                              _E_TEXT),
            (a.asset_name,                              _E_TEXT),
            (a.brand,                                   _E_MUTED),
            (a.status,                                  sc),
            (a.category.name if a.category else '',     _E_TEXT),
            (a.location.name if a.location else '',     _E_TEXT),
            (getattr(a, 'approval_status', ''),         _E_MUTED),
        ]
        for col, (val, fg) in enumerate(row_data, 1):
            _data_cell(ws2, i, col, val, bg=bg, fg=fg)
        ws2.row_dimensions[i].height = 16

    for col, w in zip('ABCDEFG', [14, 26, 16, 14, 18, 18, 14]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Repairs ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Repairs")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:F1')
    t3 = ws3.cell(row=1, column=1, value='REPAIR TICKET LOG')
    t3.font = _xfont(bold=True, size=13, color=_E_ACCENT)
    t3.fill = _xfill(_E_DARK)
    ws3.row_dimensions[1].height = 24

    r_hdrs = ['Ticket ID', 'Asset Code', 'Status', 'Repair Cost', 'Start Date', 'Completion']
    for col, h in enumerate(r_hdrs, 1):
        _header_cell(ws3, 2, col, h)
    ws3.row_dimensions[2].height = 18

    REPAIR_COLORS = {'OPEN': _E_RED, 'IN_PROGRESS': _E_YELLOW, 'COMPLETED': _E_GREEN}
    xr_qs = RepairTicket.objects.select_related('asset')
    if user and getattr(user, 'role', None) != "SUPER_ADMIN" and user.department:
        xr_qs = xr_qs.filter(asset__department=user.department)
    repairs = xr_qs.all()
    for i, r in enumerate(repairs, start=3):
        sc = REPAIR_COLORS.get(r.status, _E_TEXT)
        bg = _E_DARK if i % 2 == 0 else _E_CARD
        _data_cell(ws3, i, 1, r.id,                                                bg=bg)
        _data_cell(ws3, i, 2, r.asset.asset_code if r.asset else '',               bg=bg)
        _data_cell(ws3, i, 3, r.status,                           bg=bg, fg=sc, bold=True)
        _data_cell(ws3, i, 4, float(r.repair_cost),                                bg=bg, fg=_E_ACCENT)
        _data_cell(ws3, i, 5, r.start_date.strftime('%Y-%m-%d') if r.start_date else '', bg=bg, fg=_E_MUTED)
        _data_cell(ws3, i, 6,
                   r.completion_date.strftime('%Y-%m-%d') if r.completion_date else '',
                   bg=bg, fg=_E_MUTED)
        ws3.row_dimensions[i].height = 16

    for col, w in zip('ABCDEF', [12, 16, 15, 16, 14, 14]):
        ws3.column_dimensions[col].width = w

    # ── Sheet 4: Procurement ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Procurement")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:E1')
    t4 = ws4.cell(row=1, column=1, value='PROCUREMENT REQUESTS')
    t4.font = _xfont(bold=True, size=13, color=_E_ACCENT)
    t4.fill = _xfill(_E_DARK)
    ws4.row_dimensions[1].height = 24

    p_hdrs = ['Request ID', 'Description', 'Status', 'Created', 'Updated']
    for col, h in enumerate(p_hdrs, 1):
        _header_cell(ws4, 2, col, h)

    PROC_COLORS = {'PENDING': _E_YELLOW, 'APPROVED': _E_GREEN, 'REJECTED': _E_RED}
    xp_qs = ProcurementRequest.objects.all()
    if user and getattr(user, 'role', None) != "SUPER_ADMIN" and user.department:
        xp_qs = xp_qs.filter(department=user.department)
    procs = xp_qs.all()
    for i, p in enumerate(procs, start=3):
        sc = PROC_COLORS.get(p.approval_status, _E_TEXT)
        bg = _E_DARK if i % 2 == 0 else _E_CARD
        _data_cell(ws4, i, 1, p.id,                                                   bg=bg)
        _data_cell(ws4, i, 2, getattr(p, 'description', str(p)),                      bg=bg)
        _data_cell(ws4, i, 3, p.approval_status,                bg=bg, fg=sc, bold=True)
        _data_cell(ws4, i, 4, p.created_at.strftime('%Y-%m-%d') if p.created_at else '', bg=bg, fg=_E_MUTED)
        _data_cell(ws4, i, 5, p.updated_at.strftime('%Y-%m-%d') if hasattr(p, 'updated_at') and p.updated_at else '', bg=bg, fg=_E_MUTED)
        ws4.row_dimensions[i].height = 16

    for col, w in zip('ABCDE', [12, 36, 14, 14, 14]):
        ws4.column_dimensions[col].width = w

    # ── Sheet 5: Compliance ────────────────────────────────────────────────
    if metrics.get('total_services', 0) > 0:
        ws5 = wb.create_sheet("Compliance")
        ws5.sheet_view.showGridLines = False

        ws5.merge_cells('A1:E1')
        t5 = ws5.cell(row=1, column=1, value='SERVICE COMPLIANCE & EXPIRY')
        t5.font = _xfont(bold=True, size=13, color=_E_ACCENT)
        t5.fill = _xfill(_E_DARK)
        ws5.row_dimensions[1].height = 24

        _header_cell(ws5, 2, 1, 'Active Services')
        _header_cell(ws5, 2, 2, 'Compliance Score')
        _header_cell(ws5, 2, 3, 'Expiring (30d)')
        _header_cell(ws5, 2, 4, 'Expired')
        _data_cell(ws5, 3, 1, metrics['total_services'], bg=_E_CARD, fg=_E_ACCENT, bold=True)
        _data_cell(ws5, 3, 2, f"{metrics['compliance_score']}%", bg=_E_CARD, fg=_E_GREEN, bold=True)
        _data_cell(ws5, 3, 3, metrics['expiring_30'], bg=_E_CARD, fg=_E_YELLOW, bold=True)
        _data_cell(ws5, 3, 4, metrics['expired_services'], bg=_E_CARD, fg=_E_RED, bold=True)

        ws5.merge_cells('A5:E5')
        h5 = ws5.cell(row=5, column=1, value='SERVICES EXPIRING IN NEXT 30 DAYS')
        h5.font = _xfont(bold=True, size=11, color=_E_ACCENT)
        h5.fill = _xfill(_E_DARK)

        expiring_by_type = metrics.get('expiring_by_type', [])
        if expiring_by_type:
            ws5.merge_cells('A6:B6')
            h5t = ws5.cell(row=6, column=1, value='Breakdown by Service Type')
            h5t.font = _xfont(bold=True, size=9, color=_E_MUTED)
            h5t.fill = _xfill(_E_PANEL)
            _header_cell(ws5, 7, 1, 'Service Type')
            _header_cell(ws5, 7, 2, 'Count')
            for i, et in enumerate(expiring_by_type, start=8):
                bg = _E_DARK if i % 2 == 0 else _E_CARD
                _data_cell(ws5, i, 1, et['service_type__name'], bg=bg)
                _data_cell(ws5, i, 2, et['count'], bg=bg, fg=_E_YELLOW, bold=True)
            start_row = 8 + len(expiring_by_type) + 1
        else:
            start_row = 7

        c_hdrs = ['Asset Code', 'Asset Name', 'Service Type', 'Expiry Date', 'Days Left']
        for col, h in enumerate(c_hdrs, 1):
            _header_cell(ws5, start_row, col, h)

        exp_details = metrics.get('expiring_details', [])
        if exp_details:
            for i, ed in enumerate(exp_details, start=start_row + 1):
                exp_d = date.fromisoformat(ed['end_date']) if isinstance(ed['end_date'], str) else ed['end_date']
                dl = (exp_d - date.today()).days if exp_d else 0
                bg = _E_DARK if i % 2 == 0 else _E_CARD
                _data_cell(ws5, i, 1, ed['asset__asset_code'], bg=bg)
                _data_cell(ws5, i, 2, ed['asset__asset_name'][:30], bg=bg)
                _data_cell(ws5, i, 3, ed['service_type__name'], bg=bg)
                _data_cell(ws5, i, 4, str(exp_d), bg=bg, fg=_E_MUTED)
                fg = _E_RED if dl <= 7 else _E_YELLOW if dl <= 15 else _E_ACCENT
                _data_cell(ws5, i, 5, f"{dl}d", bg=bg, fg=fg, bold=True)
        else:
            ws5.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=5)
            nc = ws5.cell(row=start_row + 1, column=1, value='No services are expiring in the next 30 days.')
            nc.font = _xfont(size=9, color=_E_MUTED)

        for col, w in zip('ABCDE', [16, 30, 18, 14, 12]):
            ws5.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
def generate_weekly_report(user=None, report=None):
    metrics     = collect_kpi_metrics(user=user)
    charts      = generate_charts(metrics)
    ts = __import__('datetime').datetime.now().strftime('%B %d, %Y %I:%M %p')
    pdf_data    = generate_pdf(metrics, charts, title=f"Weekly Executive Report — {ts}", user=user)
    excel_data  = generate_excel(metrics, user=user)

    if report is None:
        report = Report.objects.create(
            title=f"Weekly Executive Report — {ts}",
            report_type='WEEKLY',
            generated_by=user,
            department=user.department if user and user.department else None,
            summary_data=metrics,
            is_scheduled=(user is None),
            pdf_data=pdf_data,
            excel_data=excel_data,
            chart_data=charts.get('asset_status', b''),
        )
    else:
        report.title = f"Weekly Executive Report — {ts}"
        report.report_type = 'WEEKLY'
        report.generated_by = user
        report.department = user.department if user and user.department else None
        report.summary_data = metrics
        report.is_scheduled = (user is None)
        report.pdf_data = pdf_data
        report.excel_data = excel_data
        report.chart_data = charts.get('asset_status', b'')
        report.save()
    return report